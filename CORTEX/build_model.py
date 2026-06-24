from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORANGE="C45A1A"; BLACK="1C1917"; CREAM="F5F2EE"; LGREY="EDEAE4"
HFONT=Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE=Font(name="Arial", bold=True, color=BLACK, size=16)
SUB=Font(name="Arial", bold=True, color=ORANGE, size=12)
BOLD=Font(name="Arial", bold=True, color="000000", size=10)
NORM=Font(name="Arial", color="000000", size=10)
BLUE=Font(name="Arial", color="0000FF", size=10)
GREEN=Font(name="Arial", color="008000", size=10)
hfill=PatternFill("solid", fgColor=ORANGE)
cfill=PatternFill("solid", fgColor=CREAM)
gfill=PatternFill("solid", fgColor=LGREY)
yfill=PatternFill("solid", fgColor="FFF2CC")
thin=Side(style="thin", color="C9C4BA")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center", vertical="center", wrap_text=True)
left=Alignment(horizontal="left", vertical="center", wrap_text=True)
right=Alignment(horizontal="right", vertical="center")

def hdr(ws, row, vals, start=1, fill=hfill, font=HFONT):
    for i,v in enumerate(vals):
        c=ws.cell(row=row, column=start+i, value=v); c.font=font; c.fill=fill
        c.alignment=center; c.border=border
def put(ws, r, c, v, font=NORM, align=left, fmt=None, fill=None, bd=True):
    cell=ws.cell(row=r, column=c, value=v); cell.font=font; cell.alignment=align
    if fmt: cell.number_format=fmt
    if fill: cell.fill=fill
    if bd: cell.border=border
    return cell

wb=Workbook()
ws=wb.active; ws.title="Assumptions"; ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=42
ws.column_dimensions["C"].width=18; ws.column_dimensions["D"].width=40
put(ws,2,2,"CORTEX — Financial & Pricing Model", TITLE, left, bd=False)
put(ws,3,2,"Business Intelligence Operating System (BIOS)  ·  v1.0  ·  June 2026", SUB, left, bd=False)
put(ws,4,2,"SEA/Malaysia-anchored pricing  ·  EU & Premium variants  ·  recurring figures standardised to USD", NORM, left, bd=False)
put(ws,6,2,"KEY ASSUMPTIONS (blue = editable inputs)", BOLD, left, bd=False)
hdr(ws,7,["Assumption","Value","Note"],start=2)
rows=[
 ("FX: USD per MYR",0.213,"1 USD ~ 4.70 MYR (Jun 2026)"),
 ("FX: USD per EUR",1.07,"1 EUR ~ 0.935 USD"),
 ("Blended implementation labour cost / hr (USD)",42,"Internal: Karim + KL contractors, Yr1-2"),
 ("Subscription COGS - Starter / mo (USD)",25,"Shared multi-tenant hosting + LLM tokens"),
 ("Subscription COGS - Growth / mo (USD)",90,"Multi-tenant, more docs + integrations"),
 ("Subscription COGS - Professional / mo (USD)",350,"Multi/isolated tenant, heavier compute"),
 ("Subscription COGS - Enterprise / mo (USD)",1400,"Single-tenant / VPC, heavy LLM + governance"),
 ("Gross logo retention (annual)",0.85,"Active-base roll-forward"),
 ("Net revenue retention (expansion)",1.15,"Land-and-expand up the maturity ladder"),
]
r=8
for name,val,note in rows:
    put(ws,r,2,name,NORM); c=put(ws,r,3,val,BLUE,right)
    if isinstance(val,float) and val<2: c.number_format="0.000" if val<0.3 else "0.00"
    if "COGS" in name or "labour" in name: c.number_format="$#,##0"
    put(ws,r,4,note,NORM); r+=1
for rr in range(8,8+len(rows)): ws.cell(row=rr,column=3).fill=yfill
put(ws,r+1,2,"TABS", BOLD, left, bd=False)
for i,(t,desc) in enumerate([
 ("Offer Structure","4-tier ladder: features, limits, AI, automations, support"),
 ("Pricing - SEA","Malaysia / SEA price book (MYR), anchor market"),
 ("Pricing - EU","Europe / DACH price book (EUR)"),
 ("Pricing - Premium","Global Premium Enterprise price book (USD)"),
 ("Margin Analysis","Per-tier subscription & implementation margins"),
 ("Revenue Model","3-year bookings, ARR, P&L, EBITDA"),
 ("Headcount","Hiring plan vs. revenue per phase"),
]):
    put(ws,r+2+i,2,t,BOLD,left,bd=False); put(ws,r+2+i,3,"->",NORM,center,bd=False); put(ws,r+2+i,4,desc,NORM,left,bd=False)

ws=wb.create_sheet("Offer Structure"); ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=28
for col in "BCDE": ws.column_dimensions[col].width=30
put(ws,2,1,"The CORTEX Product Ladder - 4 Tiers", TITLE, left, bd=False)
put(ws,3,1,"One platform, five maturity levels, four company sizes. Customers climb - never re-platform.", NORM, left, bd=False)
hdr(ws,5,["","STARTER","GROWTH","PROFESSIONAL","ENTERPRISE"])
specs=[
 ("Target customer","Solo / micro (1-10)","SME (10-250)","Mid-market (250-2,500)","Enterprise (2,500-10,000+)"),
 ("Maturity journey","Level 1 -> 2","Level 2 -> 3","Level 3 -> 4","Level 4 -> 5"),
 ("Outcome","Find & trust answers","Brain runs the work","Brain helps decide","Business runs on it"),
 ("Brain layers active","2-3 (Knowledge, Ops +1)","4-5","6-7","All 8 lobes"),
 ("Documents ingested","Up to 500","Up to 5,000","Up to 50,000","Unlimited"),
 ("Data sources","3 (uploaded)","8 (docs + first APIs)","15+ live integrations","Unlimited + custom connectors"),
 ("AI functions","Grounded Q&A, source citation, confidence bands","+ On-brand content, ops drafting, task running","+ Decision scoring engine, agents, scenarios","+ Custom scoring models, multi-brain, full agents"),
 ("Automations","Read-only briefs only","Basic schedules & reminders","Workflows + tool integrations","Full autonomous cadence + custom triggers"),
 ("Dashboard","Search + ask box","+ Knowledge-graph view, briefs","+ Scorecards, reports, graph analytics","+ Role-aware enterprise dashboards"),
 ("Governance / RBAC","Owner / flat","A few roles","Department clearances","Fine-grained, SSO, full audit trail"),
 ("Deployment","Shared multi-tenant","Multi-tenant","Multi-tenant or isolated","Single-tenant / VPC / on-prem"),
 ("Trust & verification","Sourcing + no-fabrication","+ Human-approval tiers","+ Adversarial self-check","+ Premium governance & compliance pack"),
 ("Support","Self-serve + docs","Light onboarding","Guided implementation","Dedicated success manager + SLA"),
]
r=6
for row in specs:
    put(ws,r,1,row[0],BOLD,left,fill=gfill)
    for i in range(1,5): put(ws,r,1+i,row[i],NORM,left)
    ws.row_dimensions[r].height=42; r+=1

def pricing_sheet(name, cur, title, setup, monthly, note, addons):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=30
    for col in "BCDE": ws.column_dimensions[col].width=20
    put(ws,2,1,title, TITLE, left, bd=False); put(ws,3,1,note, NORM, left, bd=False)
    hdr(ws,5,["","STARTER","GROWTH","PROFESSIONAL","ENTERPRISE"])
    fmt=f'"{cur} "#,##0'
    put(ws,6,1,"One-time setup (CORTEX Method)",BOLD,left,fill=gfill)
    for i,v in enumerate(setup): put(ws,6,2+i,v,BLUE,right,fmt)
    put(ws,7,1,"Monthly platform fee",BOLD,left,fill=gfill)
    for i,v in enumerate(monthly): put(ws,7,2+i,v,BLUE,right,fmt)
    put(ws,8,1,"Annual platform value (12 mo)",BOLD,left,fill=gfill)
    for i in range(4):
        col=get_column_letter(2+i); put(ws,8,2+i,f"={col}7*12",NORM,right,fmt)
    put(ws,9,1,"Year-1 total (setup + 12 mo)",BOLD,left,fill=gfill)
    for i in range(4):
        col=get_column_letter(2+i); put(ws,9,2+i,f"={col}6+{col}8",BOLD,right,fmt)
    put(ws,11,1,"ADD-ONS (monetise the variable layer)",SUB,left,bd=False)
    hdr(ws,12,["Add-on","Price"]); rr=13
    for a,p in addons:
        put(ws,rr,1,a,NORM); put(ws,rr,2,p,BLUE,right); rr+=1

pricing_sheet("Pricing - SEA","RM",
  "Pricing - Malaysia / Southeast Asia (MYR)  ·  ANCHOR MARKET",
  [6000,18000,45000,120000],[1200,3500,8000,18000],
  "Anchored to KL/SEA willingness-to-pay. Enterprise = annual contract; figures are floors.",
  [("Extra industry template","RM 4,000"),("Additional brand-brain (agencies)","RM 1,500/mo"),
   ("Premium integration (per system)","RM 800/mo"),("On-site capture workshop (1 day)","RM 3,500"),
   ("Compliance pack (PDPA / industry)","RM 6,000")])
pricing_sheet("Pricing - EU","EUR",
  "Pricing - Europe / DACH (EUR)",
  [2500,8000,22000,65000],[450,1400,3500,8500],
  "Higher willingness-to-pay; GDPR governance is a premium selling point, not a cost.",
  [("Extra industry template","EUR 1,500"),("Additional brand-brain (agencies)","EUR 500/mo"),
   ("Premium integration (per system)","EUR 250/mo"),("On-site capture workshop (1 day)","EUR 1,800"),
   ("Compliance pack (GDPR / industry)","EUR 2,500")])
pricing_sheet("Pricing - Premium","$",
  "Pricing - Global Premium Enterprise (USD)  ·  single-tenant / VPC",
  [3000,12000,35000,90000],[600,2000,6000,15000],
  "For regulated & large global buyers needing isolation, custom templates, dedicated success.",
  [("Extra industry template","$ 5,000"),("Additional brand-brain","$ 1,000/mo"),
   ("Premium integration (per system)","$ 500/mo"),("Dedicated success manager","$ 3,000/mo"),
   ("On-prem / VPC deployment","from $ 25,000")])

ws=wb.create_sheet("Margin Analysis"); ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=40
for col in "BCDE": ws.column_dimensions[col].width=18
put(ws,2,1,"Margin Analysis (per tier, USD)", TITLE, left, bd=False)
put(ws,3,1,"SEA prices converted to USD at the FX input. Subscription = high-margin SaaS; implementation = lower-margin services that fund the build and lock in the platform.", NORM, left, bd=False)
hdr(ws,5,["","STARTER","GROWTH","PROFESSIONAL","ENTERPRISE"])
sea_setup=[6000,18000,45000,120000]; sea_mo=[1200,3500,8000,18000]; labor_hours=[16,50,130,320]
fx="Assumptions!$C$8"; lab="Assumptions!$C$10"
cogs_refs=["Assumptions!$C$11","Assumptions!$C$12","Assumptions!$C$13","Assumptions!$C$14"]
def rl(r,label,fill=None,font=BOLD): put(ws,r,1,label,font,left,fill=fill or gfill)
rl(6,"Monthly fee (USD)")
for i in range(4): put(ws,6,2+i,f"={sea_mo[i]}*{fx}",NORM,right,"$#,##0")
rl(7,"Monthly subscription COGS (USD)")
for i in range(4): put(ws,7,2+i,f"={cogs_refs[i]}",GREEN,right,"$#,##0")
rl(8,"Monthly gross profit (USD)")
for i in range(4):
    col=get_column_letter(2+i); put(ws,8,2+i,f"={col}6-{col}7",NORM,right,"$#,##0")
rl(9,"Subscription gross margin %")
for i in range(4):
    col=get_column_letter(2+i); c=put(ws,9,2+i,f"={col}8/{col}6",BOLD,right,"0.0%"); c.fill=cfill
rl(11,"Setup price (USD)")
for i in range(4): put(ws,11,2+i,f"={sea_setup[i]}*{fx}",NORM,right,"$#,##0")
rl(12,"Implementation labour hours")
for i in range(4): put(ws,12,2+i,labor_hours[i],BLUE,right,"#,##0")
rl(13,"Implementation cost (USD)")
for i in range(4):
    col=get_column_letter(2+i); put(ws,13,2+i,f"={col}12*{lab}",NORM,right,"$#,##0")
rl(14,"Implementation gross profit (USD)")
for i in range(4):
    col=get_column_letter(2+i); put(ws,14,2+i,f"={col}11-{col}13",NORM,right,"$#,##0")
rl(15,"Implementation gross margin %")
for i in range(4):
    col=get_column_letter(2+i); c=put(ws,15,2+i,f"={col}14/{col}11",BOLD,right,"0.0%"); c.fill=cfill
rl(17,"Year-1 revenue (setup + 12mo, USD)")
for i in range(4):
    col=get_column_letter(2+i); put(ws,17,2+i,f"={col}11+{col}6*12",NORM,right,"$#,##0")
rl(18,"Year-1 COGS (impl + 12mo subs, USD)")
for i in range(4):
    col=get_column_letter(2+i); put(ws,18,2+i,f"={col}13+{col}7*12",NORM,right,"$#,##0")
rl(19,"Year-1 blended gross margin %")
for i in range(4):
    col=get_column_letter(2+i); c=put(ws,19,2+i,f"=({col}17-{col}18)/{col}17",BOLD,right,"0.0%"); c.fill=yfill
put(ws,21,1,"Scalability: as templates mature and self-serve onboarding replaces hands-on capture, implementation hours fall and subscription mix rises - blended margin trends toward 80%+.",NORM,left,bd=False)

ws=wb.create_sheet("Revenue Model"); ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=40
for col in "BCD": ws.column_dimensions[col].width=18
put(ws,2,1,"3-Year Revenue Model (USD) - Base Case", TITLE, left, bd=False)
put(ws,3,1,"Phase 1 services-led pilots -> Phase 2 productised platform -> Phase 3 platform-led growth. Active base rolls forward at 85% logo retention.", NORM, left, bd=False)
hdr(ws,5,["","Year 1 (2026-27)","Year 2 (2027-28)","Year 3 (2028-29)"])
setup_usd=[round(x*0.213) for x in sea_setup]
subs_yr=[round(x*0.213*12) for x in sea_mo]
new={"Starter":[5,30,110],"Growth":[8,22,45],"Professional":[3,8,18],"Enterprise":[0,2,6]}
tiers=["Starter","Growth","Professional","Enterprise"]
put(ws,6,1,"NEW LOGOS",SUB,left,bd=False)
r=7
for ti,t in enumerate(tiers):
    put(ws,r,1,f"  {t}",NORM)
    for y in range(3): put(ws,r,2+y,new[t][y],BLUE,right,"#,##0")
    r+=1
put(ws,r,1,"Total new logos",BOLD,left,fill=gfill)
for y in range(3):
    col=get_column_letter(2+y); put(ws,r,2+y,f"=SUM({col}7:{col}{r-1})",BOLD,right,"#,##0")
r+=2
put(ws,r,1,"ACTIVE BASE (end of year)",SUB,left,bd=False); r+=1
active_start=r; ret="Assumptions!$C$15"
for ti,t in enumerate(tiers):
    put(ws,r,1,f"  {t}",NORM); newrow=7+ti
    put(ws,r,2,f"=B{newrow}",NORM,right,"#,##0")
    put(ws,r,3,f"=B{r}*{ret}+C{newrow}",NORM,right,"#,##0")
    put(ws,r,4,f"=C{r}*{ret}+D{newrow}",NORM,right,"#,##0")
    r+=1
put(ws,r,1,"Total active clients",BOLD,left,fill=gfill)
for y in range(3):
    col=get_column_letter(2+y); put(ws,r,2+y,f"=SUM({col}{active_start}:{col}{active_start+3})",BOLD,right,"#,##0")
r+=2
put(ws,r,1,"REVENUE (USD)",SUB,left,bd=False); r+=1
setup_rev_row=r
put(ws,r,1,"Setup / implementation revenue",NORM)
for y in range(3):
    col=get_column_letter(2+y)
    parts="+".join([f"{col}{7+ti}*{setup_usd[ti]}" for ti in range(4)])
    put(ws,r,2+y,f"={parts}",NORM,right,"$#,##0")
r+=1
arr_row=r
put(ws,r,1,"Recurring revenue (from active base)",NORM)
for y in range(3):
    col=get_column_letter(2+y)
    parts="+".join([f"{col}{active_start+ti}*{subs_yr[ti]}" for ti in range(4)])
    put(ws,r,2+y,f"={parts}",NORM,right,"$#,##0")
r+=1
totrev_row=r
put(ws,r,1,"Total revenue",BOLD,left,fill=gfill)
for y in range(3):
    col=get_column_letter(2+y); put(ws,r,2+y,f"={col}{setup_rev_row}+{col}{arr_row}",BOLD,right,"$#,##0")
r+=2
put(ws,r,1,"COST & PROFIT (USD)",SUB,left,bd=False); r+=1
cogs_row=r
put(ws,r,1,"COGS (impl labour + hosting/LLM)",NORM)
for y in range(3):
    col=get_column_letter(2+y); put(ws,r,2+y,f"={col}{setup_rev_row}*0.45+{col}{arr_row}*0.2",NORM,right,"$#,##0")
r+=1
gp_row=r
put(ws,r,1,"Gross profit",BOLD,left,fill=gfill)
for y in range(3):
    col=get_column_letter(2+y); put(ws,r,2+y,f"={col}{totrev_row}-{col}{cogs_row}",BOLD,right,"$#,##0")
r+=1
put(ws,r,1,"Gross margin %",BOLD,left)
for y in range(3):
    col=get_column_letter(2+y); c=put(ws,r,2+y,f"={col}{gp_row}/{col}{totrev_row}",BOLD,right,"0.0%"); c.fill=cfill
r+=2
opex_row=r
put(ws,r,1,"Operating expenses (team, S&M, G&A)",NORM)
for y,v in enumerate([180000,520000,1150000]): put(ws,r,2+y,v,BLUE,right,"$#,##0")
r+=1
ebitda_row=r
put(ws,r,1,"EBITDA",BOLD,left,fill=gfill)
for y in range(3):
    col=get_column_letter(2+y); put(ws,r,2+y,f"={col}{gp_row}-{col}{opex_row}",BOLD,right,"$#,##0")
r+=1
put(ws,r,1,"EBITDA margin %",BOLD,left)
for y in range(3):
    col=get_column_letter(2+y); c=put(ws,r,2+y,f"={col}{ebitda_row}/{col}{totrev_row}",BOLD,right,"0.0%"); c.fill=cfill
r+=2
put(ws,r,1,"Exit-rate MRR (recurring / 12)",BOLD,left)
for y in range(3):
    col=get_column_letter(2+y); put(ws,r,2+y,f"={col}{arr_row}/12",BOLD,right,"$#,##0")

ws=wb.create_sheet("Headcount"); ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=34
for col in "BCD": ws.column_dimensions[col].width=18
put(ws,2,1,"Headcount Plan", TITLE, left, bd=False)
put(ws,3,1,"Lean services-led start; hire ahead of platform and template scale. FTEs + contractors.", NORM, left, bd=False)
hdr(ws,5,["Role","Year 1","Year 2","Year 3"])
hc=[("Founder / lead (Karim)",1,1,1),("AI / full-stack engineers",1,3,6),
    ("Implementation consultants",1,3,7),("Sales / partnerships",0,2,5),
    ("Customer success",0,1,4),("Marketing / content",1,1,3),("Ops / admin / finance",0,1,2)]
r=6
for role,a,b,c in hc:
    put(ws,r,1,role,NORM)
    for i,v in enumerate([a,b,c]): put(ws,r,2+i,v,BLUE,right,"#,##0")
    r+=1
put(ws,r,1,"Total headcount",BOLD,left,fill=gfill); tot_hc=r
for i in range(3):
    col=get_column_letter(2+i); put(ws,r,2+i,f"=SUM({col}6:{col}{r-1})",BOLD,right,"#,##0")
r+=2
put(ws,r,1,"Revenue per head (USD)",BOLD,left)
for i in range(3):
    col=get_column_letter(2+i)
    put(ws,r,2+i,f"='Revenue Model'!{col}{totrev_row}/{col}{tot_hc}",GREEN,right,"$#,##0")

wb.save("/sessions/brave-zen-fermat/mnt/Winsym Coaching/CORTEX/CORTEX_Financial_Model.xlsx")
print("saved")
